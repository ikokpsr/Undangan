from flask import Flask, request, jsonify
from flask import render_template
import openpyxl
app = Flask (__name__)

@app.route('/')
@app.route('/<nama>')
def index(nama = None):
    return render_template('index.html', user = nama)

@app.route('/undangandigital/')
def undangan():
    comments = read_comments_from_excel()
    return render_template('Undangan.html', comments=comments)

@app.route('/add_comment', methods=['POST'])
def add_comment():
    name = request.form.get('name')
    comment = request.form.get('comment')
    presence = request.form.get('presence')
    
    new_comment = f'{name},{presence},{comment}'
    write_comment_to_excel(new_comment)
    
    return jsonify({'message': 'Comment added successfully'})

def read_comments_from_excel():
    try:
        workbook = openpyxl.load_workbook('comments.xlsx')
        sheet = workbook.active
        comments = [cell.value for cell in sheet['A'] if cell.value is not None]
        return [comment.split(',') for comment in comments]
    except FileNotFoundError:
        return []

def write_comment_to_excel(comment):
    workbook = openpyxl.Workbook()

    try:
        workbook = openpyxl.load_workbook('comments.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        sheet = workbook.active
        sheet.title = 'Comments'

    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=comment)

    workbook.save('comments.xlsx')

if __name__ == '__main__':
    app.run()